"""Targeted jobs.xlsx mutations that preserve dashboard sheets and formatting."""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _applications_sheet(workbook):
    return workbook["Applications"] if "Applications" in workbook.sheetnames else workbook.worksheets[0]


def _headers(sheet) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


def _extend_tables(sheet) -> None:
    if not sheet.tables:
        return
    reference = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    for table in sheet.tables.values():
        table.ref = reference


def update_job_row(path: Path, job_id: str, updates: dict[str, Any]) -> None:
    """Update one job by ID without replacing other workbook sheets."""
    workbook = load_workbook(path)
    sheet = _applications_sheet(workbook)
    headers = _headers(sheet)
    if "id" not in headers:
        raise ValueError(f"'id' column missing in {path}")

    target_row = next(
        (
            row
            for row in range(2, sheet.max_row + 1)
            if str(sheet.cell(row, headers["id"]).value).strip() == str(job_id).strip()
        ),
        None,
    )
    if target_row is None:
        raise ValueError(f"Job with ID '{job_id}' not found in {path}")

    for header, value in updates.items():
        column = headers.get(header)
        if column is None:
            column = sheet.max_column + 1
            sheet.cell(1, column, header)
            headers[header] = column
        sheet.cell(target_row, column, value)

    _extend_tables(sheet)
    workbook.save(path)


def upsert_job_row(
    path: Path,
    record: dict[str, Any],
    *,
    preserve_existing: set[str] | None = None,
) -> tuple[str, int]:
    """Add a job record or refresh its details without replacing the workbook.

    URL fields are checked before the identifier. This allows a URL supplied by
    a user to refresh a pre-existing row even when the job board changed the
    identifier exposed in its page markup.
    """
    if not str(record.get("id", "")).strip():
        raise ValueError("A job record needs a non-empty 'id'.")

    workbook = load_workbook(path)
    sheet = _applications_sheet(workbook)
    headers = _headers(sheet)
    for header in record:
        if header not in headers:
            column = sheet.max_column + 1
            sheet.cell(1, column, header)
            headers[header] = column

    target_row = None
    for match_header in ("link", "applyUrl", "id"):
        value = str(record.get(match_header, "")).strip()
        column = headers.get(match_header)
        if not value or column is None:
            continue
        target_row = next(
            (
                row
                for row in range(2, sheet.max_row + 1)
                if str(sheet.cell(row, column).value or "").strip() == value
            ),
            None,
        )
        if target_row is not None:
            break

    action = "updated" if target_row is not None else "inserted"
    if target_row is None:
        target_row = sheet.max_row + 1
        style_source_row = max(2, target_row - 1)
        for column in range(1, sheet.max_column + 1):
            source = sheet.cell(style_source_row, column)
            target = sheet.cell(target_row, column)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format

    protected_headers = preserve_existing or set()
    for header, value in record.items():
        if action == "updated" and header in protected_headers:
            continue
        sheet.cell(target_row, headers[header], value)

    _extend_tables(sheet)
    workbook.save(path)
    return action, target_row


def append_job_rows(path: Path, rows: list[dict[str, Any]]) -> int:
    """Append jobs not already present, preserving dashboard and table styling."""
    if not rows:
        return 0
    workbook = load_workbook(path)
    sheet = _applications_sheet(workbook)
    headers = _headers(sheet)
    for row in rows:
        for header in row:
            if header not in headers:
                column = sheet.max_column + 1
                sheet.cell(1, column, header)
                headers[header] = column

    id_column = headers.get("id")
    existing_ids = {
        str(sheet.cell(row, id_column).value).strip()
        for row in range(2, sheet.max_row + 1)
    } if id_column else set()

    appended = 0
    for record in rows:
        record_id = str(record.get("id", "")).strip()
        if id_column and record_id in existing_ids:
            continue
        target_row = sheet.max_row + 1
        style_source_row = max(2, target_row - 1)
        for column in range(1, sheet.max_column + 1):
            source = sheet.cell(style_source_row, column)
            target = sheet.cell(target_row, column)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
        for header, value in record.items():
            sheet.cell(target_row, headers[header], value)
        existing_ids.add(record_id)
        appended += 1

    _extend_tables(sheet)
    workbook.save(path)
    return appended
