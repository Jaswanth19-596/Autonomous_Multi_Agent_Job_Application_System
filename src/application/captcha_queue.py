"""Persistent, human-in-the-loop handling for CAPTCHA-blocked applications."""

from __future__ import annotations

import threading
from datetime import datetime, timezone
from pathlib import Path

from src.data.jobs_workbook import get_job_row, update_job_row


JOBS_FILE = Path(__file__).resolve().parents[2] / "data" / "jobs.xlsx"
NEEDS_CAPTCHA = "Needs CAPTCHA"
NOT_APPLIED = "Not Applied"
workbook_lock = threading.Lock()


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def park_for_captcha(job_id: str, *, url: str | None = None, workbook_path: Path | None = None) -> None:
    """Take a job out of the runnable queue without losing its browser session."""
    path = workbook_path or JOBS_FILE
    updates = {
        "application_status": NEEDS_CAPTCHA,
        "captcha_detected_at": _now(),
    }
    if url:
        updates["captcha_url"] = url
    with workbook_lock:
        update_job_row(path, str(job_id), updates)


def requeue_after_captcha(job_id: str, *, workbook_path: Path | None = None) -> bool:
    """Return a CAPTCHA-held job to the tail of the normal application queue."""
    path = workbook_path or JOBS_FILE
    with workbook_lock:
        row = get_job_row(path, str(job_id))
        if row is None or str(row.get("application_status") or "").strip() != NEEDS_CAPTCHA:
            return False
        update_job_row(
            path,
            str(job_id),
            {
                "application_status": NOT_APPLIED,
                # ``get_jobs`` sorts this after jobs already waiting, so a
                # user-resolved CAPTCHA never jumps the queue.
                "queue_ready_at": _now(),
                "captcha_completed_at": _now(),
            },
        )
    return True


def has_captcha_held_job(*, workbook_path: Path | None = None) -> bool:
    """Return whether Chrome must preserve an existing CAPTCHA tab."""
    path = workbook_path or JOBS_FILE
    if not path.exists():
        return False
    # The workbook is small and this check only runs before starting a job.
    from openpyxl import load_workbook

    with workbook_lock:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook["Applications"] if "Applications" in workbook.sheetnames else workbook.worksheets[0]
            header = {str(value): index for index, value in enumerate(next(sheet.iter_rows(max_row=1, values_only=True), ())) if value}
            status_index = header.get("application_status")
            if status_index is None:
                return False
            return any(
                str(row[status_index] if status_index < len(row) else "").strip() == NEEDS_CAPTCHA
                for row in sheet.iter_rows(min_row=2, values_only=True)
            )
        finally:
            workbook.close()
