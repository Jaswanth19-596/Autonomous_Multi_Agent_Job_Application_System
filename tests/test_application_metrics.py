import asyncio
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from langchain_core.tools import tool

from src.application.metrics import (
    finish_application_metrics,
    fetch_openrouter_credit_snapshot,
    identify_jobboard,
    record_tool_call,
    record_application_destination,
    start_application_metrics,
)
from src.data.jobs_workbook import update_job_row
from src.agent.nodes import tool_node
from src.agent.tools import tools_by_name


def test_identify_jobboard_normalizes_known_ats_hosts():
    assert identify_jobboard("https://acme.wd5.myworkdayjobs.com/en-US/jobs") == "workday"
    assert identify_jobboard("https://boards.greenhouse.io/acme/jobs/123") == "greenhouse"
    assert identify_jobboard("https://jobs.example.com/apply") == "jobs.example.com"
    assert identify_jobboard(None) == "unknown"


def test_destination_replaces_linkedin_source_with_actual_ats(tmp_path: Path):
    destination = tmp_path / "jobs.xlsx"
    pd.DataFrame([{"id": "job-ats"}]).to_excel(destination, index=False)
    start_application_metrics("job-ats", "https://www.linkedin.com/jobs/view/123")
    record_application_destination("https://acme.wd5.myworkdayjobs.com/en-US/job/123")

    row = finish_application_metrics(
        "completed",
        application_url="https://acme.wd5.myworkdayjobs.com/en-US/job/123",
        workbook_path=destination,
    )

    assert row["jobboard"] == "workday"
    assert row["application_url"] == "https://acme.wd5.myworkdayjobs.com/en-US/job/123"
    frame = pd.read_excel(destination)
    assert frame.loc[0, "jobboard"] == "workday"


def test_linkedin_source_is_not_saved_as_the_jobboard(tmp_path: Path):
    destination = tmp_path / "jobs.xlsx"
    pd.DataFrame([{"id": "job-linkedin"}]).to_excel(destination, index=False)
    start_application_metrics("job-linkedin", "https://www.linkedin.com/jobs/view/123")

    row = finish_application_metrics("failed", workbook_path=destination)

    assert row["jobboard"] == "unknown"


def test_application_metrics_updates_matching_workbook_row(tmp_path: Path):
    destination = tmp_path / "jobs.xlsx"
    pd.DataFrame([
        {"id": "job-123", "title": "Engineer"},
        {"id": "job-456", "title": "Designer"},
    ]).to_excel(destination, index=False)
    start_application_metrics(
        "job-123",
        "https://jobs.ashbyhq.com/acme/123",
        {"usage_usd": 10.25, "credits_remaining_usd": 4.75},
        model="openai/gpt-5.6-luna",
    )
    record_tool_call()
    record_tool_call()

    row = finish_application_metrics(
        "completed",
        company="Acme",
        title="Engineer",
        credit_snapshot={"usage_usd": 10.375, "credits_remaining_usd": 4.625},
        workbook_path=destination,
    )

    assert row is not None
    assert row["jobboard"] == "ashby"
    assert row["model"] == "openai/gpt-5.6-luna"
    assert row["application_tool_calls"] == 2
    assert row["application_duration_seconds"] >= 0
    assert row["application_cost_usd"] == 0.125

    frame = pd.read_excel(destination)
    measured = frame[frame["id"] == "job-123"].iloc[0]
    untouched = frame[frame["id"] == "job-456"].iloc[0]
    assert measured["application_tool_calls"] == 2
    assert measured["model"] == "openai/gpt-5.6-luna"
    assert measured["application_cost_usd"] == 0.125
    assert measured["openrouter_credits_before_usd"] == 4.75
    assert measured["openrouter_credits_after_usd"] == 4.625
    assert pd.isna(untouched["application_tool_calls"])


def test_metrics_update_preserves_dashboard_sheet(tmp_path: Path):
    destination = tmp_path / "jobs.xlsx"
    with pd.ExcelWriter(destination) as writer:
        pd.DataFrame([{"id": "job-1", "title": "Engineer"}]).to_excel(
            writer, sheet_name="Applications", index=False
        )
        pd.DataFrame([{"metric": "Total Jobs", "value": 1}]).to_excel(
            writer, sheet_name="Dashboard", index=False
        )

    start_application_metrics("job-1", "https://boards.greenhouse.io/acme")
    finish_application_metrics("completed", workbook_path=destination)

    workbook = load_workbook(destination, data_only=False)
    assert workbook.sheetnames == ["Applications", "Dashboard"]
    assert workbook["Dashboard"]["A2"].value == "Total Jobs"


def test_status_update_preserves_metrics_and_dashboard_sheet(tmp_path: Path):
    destination = tmp_path / "jobs.xlsx"
    with pd.ExcelWriter(destination) as writer:
        pd.DataFrame([
            {"id": "job-1", "application_status": "In Progress", "application_tool_calls": 7}
        ]).to_excel(writer, sheet_name="Applications", index=False)
        pd.DataFrame([{"metric": "Total Jobs", "value": 1}]).to_excel(
            writer, sheet_name="Dashboard", index=False
        )

    update_job_row(destination, "job-1", {"application_status": "Applied"})

    workbook = load_workbook(destination, data_only=False)
    sheet = workbook["Applications"]
    assert workbook.sheetnames == ["Applications", "Dashboard"]
    assert sheet["B2"].value == "Applied"
    assert sheet["C2"].value == 7


def test_worker_tool_execution_persists_all_five_core_metrics(tmp_path: Path):
    destination = tmp_path / "jobs.xlsx"
    pd.DataFrame([{"id": "job-metrics"}]).to_excel(destination, index=False)

    @tool("playwright_browser_navigate")
    def navigate(url: str) -> str:
        """Navigate to an application page."""
        return "Page URL: https://boards.greenhouse.io/acme/jobs/123"

    original_tools = dict(tools_by_name)
    tools_by_name.clear()
    tools_by_name[navigate.name] = navigate
    try:
        async def run_worker_tool():
            start_application_metrics(
                "job-metrics",
                "https://www.linkedin.com/jobs/view/123",
                {"usage_usd": 2.0, "credits_remaining_usd": 8.0},
                model="openai/gpt-5.6-luna",
            )
            await tool_node({
                "tool_calls": [{
                    "name": "playwright_browser_navigate",
                    "args": {"url": "https://www.linkedin.com/jobs/view/123"},
                    "id": "tool-call-1",
                }]
            })
            return finish_application_metrics(
                "completed",
                credit_snapshot={"usage_usd": 2.125, "credits_remaining_usd": 7.875},
                workbook_path=destination,
            )

        metrics = asyncio.run(run_worker_tool())
    finally:
        tools_by_name.clear()
        tools_by_name.update(original_tools)

    assert metrics is not None
    assert metrics["application_tool_calls"] == 1
    assert metrics["application_duration_seconds"] >= 0
    assert metrics["application_cost_usd"] == 0.125
    assert metrics["model"] == "openai/gpt-5.6-luna"
    assert metrics["jobboard"] == "greenhouse"

    frame = pd.read_excel(destination)
    row = frame.iloc[0]
    assert row["application_tool_calls"] == 1
    assert row["application_duration_seconds"] >= 0
    assert row["application_cost_usd"] == 0.125
    assert row["model"] == "openai/gpt-5.6-luna"
    assert row["jobboard"] == "greenhouse"


def test_finish_without_active_application_is_a_noop(tmp_path: Path):
    destination = tmp_path / "jobs.xlsx"
    assert finish_application_metrics("failed", workbook_path=destination) is None
    assert not destination.exists()


def test_openrouter_snapshot_without_key_is_unavailable(monkeypatch):
    monkeypatch.delenv("OPENROUTER_API_KEY", raising=False)
    assert fetch_openrouter_credit_snapshot() == {
        "usage_usd": None,
        "credits_remaining_usd": None,
    }
